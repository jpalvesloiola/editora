from openpyxl import load_workbook
from ..models import Livro

def importar_excel(arquivo_excel):
    # Carregar o arquivo Excel
    wb = load_workbook(arquivo_excel)
    ws = wb.active  # Seleciona a primeira aba

    # Iterar pelas linhas do Excel (ignorando o cabe?alho)
    for row in ws.iter_rows(min_row=2, values_only=True):
        (
            capa,
            titulo,
            subtitulo,
            projeto,
            organizadores,
            autores,
            tipo,
            isbn,
            ano_publicacao,
            numero_paginas,
            idioma,
            formato,
            disponibilidade,
            dimensoes,
            classificacao_tema,
            cdu_thema,
            palavras_chave,
            resumo,
            link_pdf
        ) = row

        Livro.objects.create(
            capa=capa,
            titulo=titulo,
            subtitulo=subtitulo,
            projeto=projeto,
            organizadores=organizadores,
            autores=autores,
            tipo=tipo,
            isbn=isbn,
            ano_publicacao=ano_publicacao,
            numero_paginas=numero_paginas,
            idioma=idioma,
            formato=formato,
            disponibilidade=disponibilidade,
            dimensoes=dimensoes,
            classificacao_tema=classificacao_tema, 
            cdu_thema=cdu_thema,
            palavras_chave=palavras_chave,
            resumo=resumo,
            link_pdf=link_pdf
        )